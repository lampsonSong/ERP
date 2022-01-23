# coding:utf-8
 
from flask import Flask, render_template, request, redirect, url_for, make_response,jsonify
from werkzeug.utils import secure_filename
import os
import cv2
import time
import pymysql
import mysql_config

import openpyxl
from datetime import timedelta

import paho.mqtt.client as mqtt
import threading
import json

#设置允许的文件格式
ALLOWED_EXTENSIONS = set(['xlsx'])
 
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS
 
app = Flask(__name__)
# 设置静态文件缓存过期时间
app.send_file_max_age_default = timedelta(seconds=1)

#app.config.from_object(mysql_config)

idx = 0
b_send_msg = False
msg = ''
b_show_msg = True

#conn = pymysql.connect(host=mysql_config.HOST, port=mysql_config.PORT, user=mysql_config.USERNAME, password=mysql_config.PASSWORD, db=mysql_config.DATABASE, charset='utf8')


## mqtt
def on_connect(client, userdata, flags, rc):
    print("Connected with result code : " + str(rc))

def on_message(client, userdata, receive_msg):
    global b_send_msg
    global msg
    global b_show_msg
    if b_show_msg:
        b_show_msg = False
        print(receive_msg.topic + " " + str(receive_msg.payload))
    if str(receive_msg.payload).startswith("b'machine"):
        msg = str(receive_msg.payload) + "--" + msg
        b_show_msg = True

def on_subscribe(client, obj, mid, granted_qos):
    print("Subscribed: "+str(mid)+" "+str(granted_qos))

def send_msg(client):
    global b_send_msg
    global msg
    while True:
        if b_send_msg:
            client.publish('stock', payload=msg, qos=0)
        time.sleep(1)

def listen_msg(client):
    client.loop_forever()

 
@app.route('/mysql', methods=['POST', 'GET'])
def hello():

    #cur = conn.cursor()
    #sql = "select * from workflow where id = 3"
    #cur.execute(sql)
    #results = cur.fetchall()
    #conn.close()

    #oneitem = results[0]
    #print("results : {}".format(oneitem[1]))
    return "hello world"


client = mqtt.Client("mqtt")
client.on_connect = on_connect
client.on_message = on_message
client.on_subscribe = on_subscribe

client.connect("192.168.1.5", 1883, 600)
#client.connect("192.168.1.4", 1883, 600)
client.subscribe("stock", 0)

@app.route('/', methods=['POST', 'GET'])  # 添加路由
def upload():
    global b_send_msg
    global msg
    listen_mqtt = threading.Thread(target=listen_msg, args=(client,))
    listen_mqtt.start()

    send_mqtt = threading.Thread(target=send_msg, args=(client,))
    send_mqtt.start()

    if request.method == 'POST':
        f = request.files['file']
 
        if not (f and allowed_file(f.filename)):
            return jsonify({"error": 1001, "msg": "please check file type, only xlsx is allowed"})
 
        user_input = request.form.get("name")
 
        basepath = os.path.dirname(__file__)  # 当前文件所在路径
 
        fname = str(hash(f)) + str(round(time.time() * 1000)) + ".xlsx"
        upload_path = os.path.join(basepath, 'static/files', secure_filename(fname))  # 注意：没有的文件夹一定要先创建，不然会提示没有该路径
        
        #f.save(upload_path)
        #print("-fname : ", fname)
 
        ##return render_template('upload_ok.html',userinput=user_input, img_name=fname, val1=time.time())

        #f = openpyxl.load_workbook(upload_path)
        #for sheet_name in f.sheetnames:
        #    ws = f[sheet_name]
        #    print(ws.max_column)

        msg = '{"m1":1, "m2":2, "task_num": 2, "current_id":0}'
        b_send_msg = True


 
    return render_template('index.html')
 
 
if __name__ == '__main__':
    # app.debug = True
    app.run(host='0.0.0.0', port=8090, debug=True)
